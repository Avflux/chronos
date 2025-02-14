import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import logging
import win32com.client
from typing import List, Dict
from ..database.connection import DatabaseConnection

logger = logging.getLogger(__name__)

class ExcelProcessor:
    MONTH_NAMES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}
    
    def __init__(self):
        self.log_callback = None
        self.progress_callback = None
        self.db = DatabaseConnection()

    def set_callbacks(self, log_callback=None, progress_callback=None):
        self.log_callback = log_callback
        self.progress_callback = progress_callback

    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        logger.info(message)

    def update_progress(self, current, total):
        if self.progress_callback:
            self.progress_callback(current, total)

    def convert_time_to_decimal(self, time_value) -> float:
        """
        Converte tempo para decimal, aceitando tanto string HH:MM:SS quanto timedelta
        """
        try:
            if not time_value:
                return 0.0

            # Se for timedelta, extrair horas, minutos e segundos
            if isinstance(time_value, timedelta):
                total_seconds = int(time_value.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
            # Se for string, fazer split
            elif isinstance(time_value, str):
                if time_value == "00:00:00":
                    return 0.0
                hours, minutes, seconds = map(int, time_value.split(':'))
            else:
                logger.error(f"Tipo de tempo não suportado: {type(time_value)}")
                return 0.0

            decimal_hours = hours + (minutes / 60.0) + (seconds / 3600.0)
            # Formata para ter 4 casas decimais e usa vírgula
            return f"{decimal_hours:.4f}".replace('.', ',')
            
        except Exception as e:
            logger.error(f"Erro ao converter tempo {time_value}: {e}")
            return 0.0

    def get_daily_activities(self, user_id: int) -> List[Dict]:
        """Busca atividades do dia atual do usuário específico e agrupa duplicatas"""
        try:
            query = """
                SELECT 
                    id,
                    user_id,
                    description,
                    atividade,
                    total_time,
                    updated_at
                FROM atividades 
                WHERE user_id = %s
                AND DATE(updated_at) = CURDATE()
                AND (concluido = TRUE OR pausado = TRUE)
                ORDER BY start_time ASC
            """
            
            results = self.db.execute_query(query, (user_id,))
            
            # Dicionário para agrupar atividades idênticas
            grouped_activities = {}
            
            # Processar e agrupar resultados
            for activity in results:
                # Criar uma chave única usando descrição e atividade
                key = (activity['description'], activity['atividade'])
                
                # Converter tempo total para timedelta para facilitar a soma
                if isinstance(activity['total_time'], str):
                    h, m, s = map(int, activity['total_time'].split(':'))
                    total_time = timedelta(hours=h, minutes=m, seconds=s)
                else:
                    total_time = activity['total_time']
                
                if key in grouped_activities:
                    # Se já existe, soma o tempo
                    grouped_activities[key]['total_time'] += total_time
                else:
                    # Se não existe, cria novo registro
                    grouped_activities[key] = {
                        'id': activity['id'],
                        'description': activity['description'],
                        'atividade': activity['atividade'],
                        'total_time': total_time,
                        'updated_at': activity['updated_at']
                    }
            
            # Converter resultados agrupados para lista
            processed_activities = []
            for activity in grouped_activities.values():
                # Converter tempo total para decimal
                total_time_decimal = self.convert_time_to_decimal(activity['total_time'])
                
                processed_activities.append({
                    'id': activity['id'],
                    'description': activity['description'],
                    'atividade': activity['atividade'],
                    'total_time': total_time_decimal,
                    'updated_at': activity['updated_at']
                })
            
            return processed_activities
            
        except Exception as e:
            logger.error(f"Erro ao buscar atividades: {e}")
            return []

    def process_activities_to_excel(self, user_id: int, caminho_destino: str) -> bool:
        """Processa atividades do banco para o Excel"""
        self.log("\n[INFO] Iniciando processamento de atividades para Excel...")
        
        if not os.path.exists(caminho_destino):
            self.log("[ERRO] Arquivo de destino não encontrado!")
            return False

        activities = self.get_daily_activities(user_id)
        if not activities:
            self.log("[INFO] Nenhuma atividade encontrada para hoje.")
            return True

        self.log(f"[INFO] Total de atividades encontradas: {len(activities)}")
        
        excel = None
        workbook = None

        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False

            caminho_destino = os.path.normpath(caminho_destino)
            workbook = excel.Workbooks.Open(caminho_destino)

            mes_atual = datetime.now().month
            sheet_name = self.MONTH_NAMES.get(mes_atual)
            if sheet_name:
                try:
                    sheet = workbook.Worksheets(sheet_name)
                    sheet.Select()
                except:
                    sheet = workbook.ActiveSheet
            else:
                sheet = workbook.ActiveSheet
            
            dia_atual = datetime.now().day
            coluna_destino = self.get_column_letter(dia_atual)
            
            total_activities = len(activities)

            for idx, activity in enumerate(activities, 1):
                try:
                    self.log(f"\n[DEBUG] Processando atividade {idx}")
                    self.log(f"[DEBUG] Descrição: {activity['description']}")
                    self.log(f"[DEBUG] Atividade: {activity['atividade']}")
                    
                    linha_atual = 7
                    linha_encontrada = None
                    
                    # Procurar linha apropriada
                    while linha_atual <= 101:
                        desc_excel = sheet.Range(f"B{linha_atual}").Value
                        ativ_excel = sheet.Range(f"C{linha_atual}").Value
                        
                        self.log(f"[DEBUG] Linha {linha_atual}:")
                        self.log(f"[DEBUG] Excel Descrição: {desc_excel}")
                        self.log(f"[DEBUG] Excel Atividade: {ativ_excel}")
                        
                        if desc_excel is None:
                            # Encontrou linha vazia
                            self.log("[DEBUG] Encontrou linha vazia - usando para nova entrada")
                            linha_encontrada = linha_atual
                            break
                            
                        if desc_excel == activity['description']:
                            if ativ_excel == activity['atividade']:
                                # Descrição e atividade iguais
                                self.log("[DEBUG] Encontrou descrição e atividade iguais - usando linha existente")
                                linha_encontrada = linha_atual
                                break
                            else:
                                # Descrição igual mas atividade diferente
                                self.log("[DEBUG] Encontrou descrição igual mas atividade diferente - continuando busca")
                                
                        linha_atual += 1
                    
                    if linha_encontrada is None:
                        self.log("[DEBUG] Nenhuma linha adequada encontrada")
                        continue
                        
                    # Processamento da linha encontrada
                    celula_destino = f"{coluna_destino}{linha_encontrada}"
                    
                    # Se é uma nova linha, escreve descrição e atividade
                    if desc_excel is None:
                        self.log(f"[DEBUG] Escrevendo nova entrada na linha {linha_encontrada}")
                        sheet.Range(f"B{linha_encontrada}").Value = activity['description']
                        sheet.Range(f"C{linha_encontrada}").Value = activity['atividade']
                    
                    # Calcula valor acumulado dos dias anteriores
                    valor_acumulado = 0
                    try:
                        for dia in range(1, dia_atual):
                            coluna_anterior = self.get_column_letter(dia)
                            celula_anterior = f"{coluna_anterior}{linha_encontrada}"
                            valor_anterior = sheet.Range(celula_anterior).Value
                            
                            if valor_anterior is not None:
                                self.log(f"[DEBUG] Dia {dia} - Valor anterior encontrado: {valor_anterior}")
                                if isinstance(valor_anterior, (int, float)):
                                    valor_acumulado += float(valor_anterior)
                        
                        self.log(f"[DEBUG] Valor total acumulado dos dias anteriores: {valor_acumulado}")
                    except Exception as e:
                        self.log(f"[ERRO] Falha ao calcular valores anteriores: {str(e)}")
                        valor_acumulado = 0
                    
                    # Processa o valor do tempo
                    try:
                        valor = str(activity['total_time']).replace(',', '.')
                        numero = float(valor)
                        
                        # Subtrai o valor acumulado se necessário
                        if valor_acumulado > 0:
                            numero_ajustado = numero - valor_acumulado
                            self.log(f"[DEBUG] Valor original: {numero}")
                            self.log(f"[DEBUG] Valor após subtração do acumulado: {numero_ajustado}")
                            numero = numero_ajustado
                        
                        # Verifica se a célula está vazia antes de escrever
                        if not sheet.Range(celula_destino).Value:
                            self.log(f"[DEBUG] Inserindo valor {numero} em {celula_destino}")
                            sheet.Range(celula_destino).Value = numero
                            sheet.Calculate()
                            workbook.Application.CalculateFull()
                    except ValueError as e:
                        self.log(f"[ERRO] Valor inválido para conversão: {valor}")
                        continue
                    
                    self.update_progress(idx, total_activities)
                    
                except Exception as e:
                    self.log(f"[ERRO] Falha ao processar atividade {idx}: {str(e)}")
                    continue

            workbook.Save()
            self.log("[INFO] Arquivo Excel atualizado com sucesso!")
            return True

        except Exception as e:
            self.log(f"[ERRO] Falha ao processar arquivo: {str(e)}")
            return False
            
        finally:
            try:
                if workbook:
                    workbook.Close(SaveChanges=True)
                if excel:
                    del excel
            except Exception as e:
                self.log(f"[ERRO] Falha ao fechar recursos: {str(e)}")

    def get_column_letter(self, day: int) -> str:
        """Retorna a letra da coluna baseado no dia do mês"""
        if day < 1 or day > 31:
            raise ValueError(f"Dia inválido: {day}")
        
        if day <= 23:
            return chr(ord('D') + day - 1)
        
        first_letter = 'A'
        second_letter = chr(ord('A') + (day - 24))
        return f"{first_letter}{second_letter}"